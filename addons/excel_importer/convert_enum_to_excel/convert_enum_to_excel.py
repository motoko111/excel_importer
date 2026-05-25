import json
import os
import sys
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

INPUT_DIR = "input"
INPUT_ENUM_FILE_PATH = "enum/enum_list.json"
TARGET_FILE_REGEX = ".*"

# enum値を格納する専用シート名
ENUM_SHEET_NAME = "_enums"

if len(sys.argv) >= 2 and len(sys.argv[1]) > 0:
    INPUT_DIR = sys.argv[1]
    print("inputディレクトリを設定:" + INPUT_DIR)
if len(sys.argv) >= 3 and len(sys.argv[2]) > 0:
    INPUT_ENUM_FILE_PATH = sys.argv[2]
    print("読み込むjsonファイルパスを設定:" + INPUT_ENUM_FILE_PATH)
if len(sys.argv) >= 4 and len(sys.argv[3]) > 0:
    TARGET_FILE_REGEX = sys.argv[3]
    print("対象となるファイル正規表現を設定:" + TARGET_FILE_REGEX)


def write_enum_sheet(wb, enum_map: dict[str, list[str]], needed_keys: set[str]) -> dict[str, str]:
    """
    enum定義のうち、対象シートで実際に使われているキー(needed_keys)だけを
    専用シート(_enums)に列ごとに書き込み、
    各enumキーに対応する範囲参照文字列(formula1で使う形式)を返す。

    レイアウト:
        row1: enumキー名(ヘッダー)
        row2以降: enum値(縦方向)

    Args:
        wb: 対象ワークブック
        enum_map: JSONから読み込んだenum定義全体
        needed_keys: 対象シートのヘッダー行に存在し、書き込みが必要なキーの集合

    Returns:
        dict[str, str]: { enumキー名: "'_enums'!$A$2:$A$51" のような範囲参照 }
    """
    # 既存のenumシートがあれば一旦削除して作り直す(値の追加・削除に追従するため)
    if ENUM_SHEET_NAME in wb.sheetnames:
        del wb[ENUM_SHEET_NAME]
    enum_ws = wb.create_sheet(title=ENUM_SHEET_NAME)

    ref_map: dict[str, str] = {}
    col_idx = 1
    for key, values in enum_map.items():
        # 対象シートで使われていないenumはスキップ(_enumsに書き込まない)
        if key not in needed_keys:
            continue
        if not values:
            continue

        col_letter = get_column_letter(col_idx)
        # ヘッダー(1行目)にキー名
        enum_ws.cell(row=1, column=col_idx, value=key)
        # 2行目以降に値を縦に並べる
        for i, v in enumerate(values):
            enum_ws.cell(row=2 + i, column=col_idx, value=v)

        start_cell = f"${col_letter}$2"
        end_cell = f"${col_letter}${1 + len(values)}"
        # シート名をシングルクォートで囲む(スペースや記号を含む場合に安全)
        ref_map[key] = f"'{ENUM_SHEET_NAME}'!{start_cell}:{end_cell}"
        col_idx += 1

    # enumシートは参照専用なので非表示にする
    # enum_ws.sheet_state = "hidden"
    return ref_map


def apply_enum_list_to_excel(
    excel_path: str,
    enum_json_path: str,
    target_sheet: str = "Sheet1",
    header_row: int = 1,     # enumキーを探す行
    start_row: int = 2       # リスト設定を始める行
):
    """
    Godotで出力したenum定義JSONをもとに、
    各Excelファイル内にenum専用シート(_enums)を作成し、
    対象シートのキー名と一致する列に対して、
    enumシートの範囲を参照するリスト選択を設定する。

    Args:
        excel_path (str): 編集対象のExcelファイル
        enum_json_path (str): Godotで生成したenums.json
        target_sheet (str): 対象シート名
        header_row (int): enumキーを探す行（例: 1）
        start_row (int): リスト設定を開始する行（例: 2）
    """
    #
    if not(is_target_file_path(excel_path)):
        print(f"対象ファイルの正規表現に一致しないため処理対象から除外します。:{excel_path}")
        return False
    #
    if is_file_locked(excel_path):
         print(f"エラー!!!:{excel_path}のファイルが開かれています。")
         return False

    # JSON読み込み
    with open(enum_json_path, "r", encoding="utf-8") as f:
        enum_map: dict[str, list[str]] = json.load(f)

    # Excel読み込み
    wb = load_workbook(excel_path)
    if target_sheet not in wb.sheetnames:
        print(f"エラー!!!:指定されたシート '{target_sheet}' が存在しません。")
        return False
    ws = wb[target_sheet]

    max_col = ws.max_column
    max_row = ws.max_row
    # データ行が少なくstart_rowを下回る場合は、最低でもstart_rowまで範囲を確保
    if max_row < start_row:
        max_row = start_row

    # 先に対象シートのヘッダー行を走査し、実際に使われているenumキーを収集する。
    # こうすることで、_enumsには必要なenumだけを書き込める。
    needed_keys: set[str] = set()
    for col in range(1, max_col + 1):
        key = ws.cell(row=header_row, column=col).value
        if not key:
            continue
        key = str(key).strip()
        if key in enum_map:
            needed_keys.add(key)

    # 対象シートで使われているenumだけを専用シートに書き込み、範囲参照マップを取得
    ref_map = write_enum_sheet(wb, enum_map, needed_keys)

    updated = 0

    # キー探索行の全セルをチェック
    for col in range(1, max_col + 1):
        key = ws.cell(row=header_row, column=col).value
        if not key:
            continue

        key = str(key).strip()
        if key in ref_map:
            formula = ref_map[key]

            # リスト適用範囲を設定
            start = ws.cell(row=start_row, column=col).coordinate
            end = ws.cell(row=max_row, column=col).coordinate

            # 削除したい範囲を含む入力規則だけを除外
            new_validations = []
            for dv in ws.data_validations.dataValidation:
                # この範囲を削除したい場合（例: A1:A10）
                if f"{start}:{end}" in dv.sqref:
                    continue  # 削除
                new_validations.append(dv)

            # フィルタリング後の入力規則を再設定
            ws.data_validations.dataValidation = new_validations

            # enumシートの範囲を参照するDataValidationを設定
            # formula1に範囲参照を指定することで255文字制限を回避
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(dv)

            dv.add(f"{start}:{end}")

            print(f"{key} -> {start}:{end} にリスト設定(参照: {formula})")
            updated += 1

    # enumシートは書き込んだので、リスト適用が無くても保存する
    try:
        wb.save(excel_path)
    except ValueError as e:
        print(f"エラー!!!:{excel_path}の保存時にエラーが発生しました。ファイルが開かれている場合は閉じてください。\n: {e}")
        return False
    print(f"完了: {excel_path} に {updated} 列のリスト選択を設定しました。")
    return True

def is_file_locked(filepath: str) -> bool:
    """ファイルが他のプロセス（Excelなど）で開かれているか確認"""
    if not os.path.exists(filepath):
        return False

    try:
        # 'a' モードで開いてすぐ閉じる
        with open(filepath, 'a'):
            pass
        return False
    except PermissionError:
        return True

def is_target_file_path(filepath: str) -> bool:
    """ 対象ファイル名が条件に一致しているかチェック """
    return re.match(TARGET_FILE_REGEX, os.path.basename(filepath))

def apply_enum_list_to_excel_dir(dir_path, enum_json_path):
    files = get_input_files(dir_path)
    for file in files:
        filename = os.path.splitext(os.path.basename(file))[0]
        print("================== " + filename + " の処理を開始 ==================")
        apply_enum_list_to_excel(
        excel_path=file,
        enum_json_path=enum_json_path,
        target_sheet=filename,
        header_row=3,   # header_row行目でキーを検索
        start_row=5     # start_row行目以降にリスト設定
        )
        print("================== " + filename + " の処理を終了 ==================")

def get_input_files(dir_path):
    '''
    input配下のファイルを全て取得
    '''
    files = os.listdir(dir_path)
    files_file = [f for f in files if os.path.isfile(os.path.join(dir_path, f))]
    ret = []
    for f in files_file:
        if "~" in f:
            continue
        if not(".xlsx" in f) :
            continue
        ret.append(dir_path + "/" + f)
    return ret

def main():
    apply_enum_list_to_excel_dir(INPUT_DIR, INPUT_ENUM_FILE_PATH)

main()
