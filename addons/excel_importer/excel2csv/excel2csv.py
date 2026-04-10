import re
import sys
import os
import openpyxl
import json
import shutil

DELIMITER_STR = ","
LINE_STR = "\n"
INPUT_DIR = "input"
OUTPUT_DIR = "output"

if len(sys.argv) >= 2 and len(sys.argv[1]) > 0:
    INPUT_DIR = sys.argv[1]
    print("inputディレクトリを設定:" + INPUT_DIR)
if len(sys.argv) >= 3 and len(sys.argv[2]) > 0:
    OUTPUT_DIR = sys.argv[2]
    print("outputディレクトリを設定:" + OUTPUT_DIR)

class DataTable:
    '''
    1シート分のデータ構造情報
    '''
    def __init__(self, name):
        self.name = name
        self.fields = []
        self.records = []
        self.fieldNameRow = -1
        self.typeRow = -1
        self.infoRow = -1
        self.startRow = -1
        self.endRow = -1
        self.isEnum = name == "Enum"

    def add_field(self,field):
        '''
        フィールド情報追加
        '''
        self.fields.append(field)

    def add_record(self, record):
        '''
        レコード追加
        '''
        self.records.append(record)

    def find_record(self, field, value):
        for record in self.records:
            if record.get_value(field) == value:
               return record
        return None

class DataField:
    '''
    1列分のデータ構造情報
    '''
    def __init__(self):
        self.type = ''
        self.name = ''
        self.info = ''
        self.column = -1

class DataRecord:
    '''
    1行分のデータ
    '''
    def __init__(self):
        self.data = {}

    def set_value(self, field, value):
        self.data[field] = value

    def get_value(self, field):
        return self.data[field]
    
    def get_export_str(self, fieldType, field, enum_split_str):
        v = self.get_value(field)
        if fieldType == "bool":
            if v == "False" or v == "FALSE" or v == False:
                return "false"
            elif v == "True" or v == "TRUE" or v == True:
                return "true"
            elif v is None:
                return "false"
            else:
                return str(v)
        elif fieldType == "string":
            if v is None:
                return '""'
            else:
                return '"' + str(v) + '"'
        elif fieldType == "name":
            if v is None:
                return '""'
            else:
                return '"' + str(v) + '"'
        elif fieldType == "int":
            if v is None:
                return "0"
            return str(v)
        elif fieldType == "float":
            if v is None:
                return "0"
            return str(v)
        elif fieldType == "double":
            if v is None:
                return "0"
            return str(v)
        elif fieldType == "id":
            return str(v)
        else:
            # enum
            if v is None:
                return ""
            else:
                # 区切りの文字はエンジン依存
                # print("fieldType:" + str(fieldType))
                # print("enum_split_str:" + str(enum_split_str))
                # print("v:" + str(v))
                return fieldType + enum_split_str + str(v)
        return str(v)

class DataEnum:
    '''
    Enumのデータ構造情報
    '''
    def __init__(self, type, type_comment, size):
        self.type = type
        self.type_comment = type_comment
        if self.type_comment == None or self.type_comment == "":
            self.type_comment = str(type)
        self.size = size
        if self.size == None or self.size == "":
            self.size = "uint8"
        self.data = []
        self.last_value = -1

    def add_data(self, name, value, text):
        if value == None or value == "":
            value = self.last_value + 1
        else :
            value = int(value)
        self.last_value = value
        self.data.append(DataEnumValue(name,value,text))

class DataEnumValue:
    '''
    Enumのデータ1個分の構造情報
    '''
    def __init__(self, name, value, text):
        self.name = name
        self.value = value
        if text is None:
            text = ""
        self.text = text

def load_excel( filepath ):
    '''
    xls/xlsmを読み込む
    '''
    wb = openpyxl.load_workbook(filepath, keep_vba=True,data_only=True)
    return wb

def load_and_analys(filepath, sheetName) -> DataTable:
    '''
    指定シートをデータテーブル化
    '''
    wb = load_excel(filepath)
    ws = wb[sheetName]
    table = analys_sheet(ws)
    return table

def load_and_analys_all(filepath):
    '''
    '''
    wb = load_excel(filepath)
    tables = analys_sheet_all(wb)
    return tables

def analys_sheet_all(wb):
    '''
    各sheetを読み込む
    '''
    tables = []
    for sheetName in wb.sheetnames:
        ws = wb[sheetName]
        tables.append(analys_sheet(ws))
    return tables

def analys_sheet(ws) -> DataTable:
    '''
    1つのsheetを読み込む
    '''
    table = DataTable( ws.title )

    # 1列目を読み込み どの行になんの項目があるかを確認
    for iRow in range(1, ws.max_row + 1):
        cell = ws.cell(iRow, 1)
        #print(u"1行目解析:" + str(cell.value))
        if cell.value == "FieldName":
            table.fieldNameRow = iRow
        elif cell.value == "Type":
            table.typeRow = iRow
        elif cell.value == "Info":
            table.infoRow = iRow
        if table.fieldNameRow != -1 and table.typeRow != -1 and table.infoRow  != -1:
            table.startRow = iRow + 1
            table.endRow = ws.max_row
            break
    
    # 各項目を登録
    for iCol in range(2, ws.max_column + 1):
        cell = ws.cell(1, iCol)
        if cell.value == "#":
            continue
        field = DataField()
        field.column = iCol
        if table.fieldNameRow != -1:
            field.name = ws.cell(table.fieldNameRow, iCol).value
        if table.typeRow != -1:
            field.type = ws.cell(table.typeRow, iCol).value
        if table.infoRow != -1:
            field.info = ws.cell(table.infoRow, iCol).value
        if field.name == None:
            continue
        print(u"項目:" + str(field.name))
        table.add_field(field)

    print(u"データ範囲:" + str(table.startRow) + "～" + str(table.endRow))
    # 各レコードを登録
    for iRow in range(table.startRow, table.endRow + 1):
        if ws.cell(iRow, 1).value == "#":
            continue
        record = DataRecord()
        for field in table.fields:
            #print("" + str(ws.cell(iRow, field.column).value))
            value = ws.cell(iRow, field.column).value
            record.set_value(field.name, value)
        table.add_record(record)

    return table

def export_enum_ue4_header(table: DataTable, setting : dict):
    '''
    Enumの出力
    '''

    path_setting = setting["UE"]["path"]

    enums = {}
    for record in table.records:
        enum_header = record.get_value("Header")
        if enum_header == None or enum_header == "":
            # データ登録
            data_enum = enums[enum_type]
            enum_name = record.get_value("Name")
            enum_value = record.get_value("Value")
            enum_text = record.get_value("Text")
            data_enum.add_data(enum_name,enum_value,enum_text)
        else:
            enum_type = record.get_value("EnumTypeName")
            enum_type_comment = record.get_value("EnumTypeNameComment")
            enum_size= record.get_value("Size")
            #print(str(enum_type) + "," + str(enum_type_comment) + "," + str(enum_size))
            # ヘッダ登録
            enums[enum_type] = DataEnum(enum_type, enum_type_comment, enum_size)
                

    txt_enums = ""
    for key in enums.keys():
        data_enum = enums[key]
        txt_enums += get_ue4_enum_str(data_enum)

    replace_map = {
        "ENUMS":txt_enums
    }

    output_dir = path_setting["EnumOutputPath"]
    if not output_dir:
        output_dir = ""
    output_dir = OUTPUT_DIR + "/" + output_dir
    os.makedirs(output_dir, exist_ok=True)
    template_file = path_setting["EnumTemplatePath"]
    output_file = output_dir + "/" + "MasterDefines" + ".h"
    export_from_template(template_file, output_file, replace_map)

def get_ue4_enum_str(data_enum : DataEnum):
    '''
    Enum1つ分のテキストを作成
    '''
    txt = ""
    txt += "// " + data_enum.type_comment + LINE_STR
    txt += "" + "UENUM(BlueprintType)" + LINE_STR
    txt += "enum class " + data_enum.type + " :" +  data_enum.size + LINE_STR
    txt += "{" + LINE_STR
    for value in data_enum.data:
        txt += "	" + value.name + " = " + str(value.value)  + "," + " // " + value.text + LINE_STR
    txt += "};" + LINE_STR
    txt += LINE_STR
    return txt

def export_enum_godot_header(table: DataTable, setting : dict):
    '''
    Enumの出力
    '''

    path_setting = setting["Godot"]["path"]

    enums = {}
    for record in table.records:
        enum_header = record.get_value("Header")
        if enum_header == None or enum_header == "":
            # データ登録
            data_enum = enums[enum_type]
            enum_name = record.get_value("Name")
            enum_value = record.get_value("Value")
            enum_text = record.get_value("Text")
            data_enum.add_data(enum_name,enum_value,enum_text)
        else:
            enum_type = record.get_value("EnumTypeName")
            enum_type_comment = record.get_value("EnumTypeNameComment")
            enum_size= record.get_value("Size")
            #print(str(enum_type) + "," + str(enum_type_comment) + "," + str(enum_size))
            # ヘッダ登録
            enums[enum_type] = DataEnum(enum_type, enum_type_comment, enum_size)
                

    txt_enums = ""
    for key in enums.keys():
        data_enum = enums[key]
        txt_enums += get_godot_enum_str(data_enum)

    replace_map = {
        "ENUMS":txt_enums
    }

    output_dir = path_setting["EnumOutputPath"]
    if not output_dir:
        output_dir = ""
    output_dir = OUTPUT_DIR + "/" + output_dir
    os.makedirs(output_dir, exist_ok=True)
    template_file = path_setting["EnumTemplatePath"]
    output_file = output_dir + "/" + "MasterDefines" + ".gd"
    export_from_template(template_file, output_file, replace_map)

def get_godot_enum_str(data_enum : DataEnum):
    '''
    Enum1つ分のテキストを作成
    '''
    txt = ""
    txt += "# " + data_enum.type_comment + LINE_STR
    txt += "enum " + data_enum.type + LINE_STR
    txt += "{" + LINE_STR
    for value in data_enum.data:
        txt += "	" + value.name + " = " + str(value.value)  + "," + " ## " + value.text + LINE_STR
    txt += "}" + LINE_STR
    txt += LINE_STR
    return txt


def export_csv(table : DataTable, setting : dict):
    '''
    DataTableをcsvに出力
    '''

    path_setting = setting["csv"]["path"]

    txt = ""
    for i in range(len(table.fields)):
        field = table.fields[i]
        if ( i >= 1):
            txt += DELIMITER_STR
        txt += field.name
    txt += LINE_STR
    for record in table.records:
        for i in range(len(table.fields)):
            field = table.fields[i]
            if ( i >= 1):
                txt += DELIMITER_STR
            value_str = record.get_export_str(field.type, field.name, ".")
            txt += value_str
            #print(u"レコード追加:" + str(field.name) + "," + str(field.type) + "," + value_str)
        txt += LINE_STR
    
    output_dir = path_setting["DataOutputPath"]
    if not output_dir:
        output_dir = ""
    output_dir = OUTPUT_DIR + "/" + output_dir
    os.makedirs(output_dir, exist_ok=True)
    if ".csv" in table.name:
        output_file = output_dir + "/" + table.name
    else:
        output_file = output_dir + "/" + table.name + ".csv"
    with open(output_file,'w',encoding='utf_8_sig', newline="") as f:
        f.write(txt)

        
def export_lua(table : DataTable, setting : dict):
    '''
    DataTableをluaに出力
    '''

    path_setting = setting["Lua"]["path"]

    TAB_STR = "	"
    txt = ""
    txt += "return { " + LINE_STR
    txt += TAB_STR + "dataName = " + '"' + table.name + '"' + "," + LINE_STR
    txt += TAB_STR + "fields = {" + LINE_STR
    for field in table.fields:
        info_str = ""
        if field.info:
            info_str = field.info.replace("\r\n","\n").replace("\r","\n").replace("\n","\\n")
        txt += TAB_STR + TAB_STR + "" + field.name + " = " + '"' + info_str + '"' + "," + LINE_STR
    txt += TAB_STR + "}," + LINE_STR # end fields
    txt += TAB_STR + "records = {" + LINE_STR
    for record in table.records:
        txt += TAB_STR + TAB_STR + "{" + LINE_STR
        for i in range(len(table.fields)):
            field = table.fields[i]
            value_str = (record.get_export_str(field.type, field.name, ".")).replace("\r\n","\n").replace("\r","\n").replace("\n","\\n")
            info_str = ""
            if field.info:
                info_str = field.info.replace("\r\n","\n").replace("\r","\n").replace("\n"," ")
            txt += TAB_STR + TAB_STR + TAB_STR + field.name + " = " + value_str + "," + "-- " + info_str + " " + LINE_STR
        txt += TAB_STR + TAB_STR + "}, " + LINE_STR
    txt += TAB_STR + "}" + LINE_STR # end records
    txt += "}" # end data
    
    output_dir = path_setting["DataOutputPath"]
    if not output_dir:
        output_dir = ""
    output_dir = OUTPUT_DIR + "/" + output_dir
    os.makedirs(output_dir, exist_ok=True)
    output_file = output_dir + "/" + table.name + ".lua"
    with open(output_file,'w',encoding='utf_8', newline="") as f:
        f.write(txt)

def export_godot_dic(table : DataTable, setting : dict):
    '''
    DataTableをgdscriptに出力
    '''

    path_setting = setting["Godot"]["path"]

    template = """
class_name ${DATA_NAME}Manager
extends MasterDataManager
static var s_instance = null
static func get_instance() -> ${DATA_NAME}Manager:
	if !s_instance:
		s_instance = ${DATA_NAME}Manager.new()
	return s_instance
func _init():
	data_name = "${DATA_NAME}"
${FIELDS}
${RECORDS}
	super._init()
"""
    TAB_STR = "	"

    fields_txt = ""
    fields_txt += TAB_STR + "fields = {" + LINE_STR
    for field in table.fields:
        info_str = ""
        if field.info:
            info_str = field.info.replace("\r\n","\n").replace("\r","\n").replace("\n","\\n")
        fields_txt += TAB_STR + TAB_STR + "" + field.name + " = " + '"' + info_str + '"' + "," + LINE_STR
    fields_txt += TAB_STR + "}" # end fields

    records_txt = ""
    records_txt += TAB_STR + "records = [" + LINE_STR
    for record in table.records:
        records_txt += TAB_STR + TAB_STR + "{" + LINE_STR
        for i in range(len(table.fields)):
            field = table.fields[i]
            str_prefix = ""
            if field.type == "name":
                str_prefix = "&"
            value_str = (str_prefix + record.get_export_str(field.type, field.name, ".")).replace("\r\n","\n").replace("\r","\n").replace("\n","\\n")
            info_str = ""
            if field.info:
                info_str = field.info.replace("\r\n","\n").replace("\r","\n").replace("\n"," ")
            records_txt += TAB_STR + TAB_STR + TAB_STR + field.name + " = " + value_str + "," + "# " + info_str + " " + LINE_STR
        records_txt += TAB_STR + TAB_STR + "}, " + LINE_STR
    records_txt += TAB_STR + "]" # end records

    txt = template.replace("${DATA_NAME}", table.name).replace("${FIELDS}", fields_txt).replace("${RECORDS}", records_txt)
    
    output_dir = path_setting["DataOutputPath"]
    if not output_dir:
        output_dir = ""
    output_dir = OUTPUT_DIR + "/" + output_dir
    os.makedirs(output_dir, exist_ok=True)
    output_file = output_dir + "/" + table.name + ".gd"
    with open(output_file,'w',encoding='utf_8', newline="") as f:
        f.write(txt)

def export_ue4_data_table(table : DataTable, setting : dict):
    '''
    DataTableをUE4のC++ファイルとして出力
    '''
    field_setting = setting["UE"]["fields"]
    path_setting = setting["UE"]["path"]

    replace_field_map = {}
    for key in field_setting:
        type_str = key
        code_type_str = field_setting[key]
        replace_field_map[type_str] = code_type_str

    txt_fields = ""
    for i in range(len(table.fields)):
        field = table.fields[i]
        txt_fields += get_ue4_field_str(field, replace_field_map)
    
    replace_map = {
        "DATA_NAME":table.name,
        "FIELDS":txt_fields
    }

    output_dir = path_setting["CodeOutputPath"]
    if not output_dir:
        output_dir = ""
    output_dir = OUTPUT_DIR + "/" + output_dir
    os.makedirs(output_dir, exist_ok=True)
    template_file = path_setting["TemplatePath"]
    output_file = output_dir + "/" + table.name + "Manager" + ".h"
    export_from_template(template_file, output_file, replace_map)

def get_ue4_field_str(field : DataField, replace_field_map):
    '''
    Field1つ分のテキストを作成
    '''
    txt = ""
    txt += "	" + "UPROPERTY(EditAnywhere)" + LINE_STR
    if field.type in replace_field_map:
        txt += "	" + replace_field_map[field.type]
    else:
        txt += "	" + field.type
    txt += " " + field.name + ";"
    txt += LINE_STR
    return txt

def export_unity_data_table(table : DataTable):
    '''
    DataTableをUnityのC#ファイルとして出力
    '''
    # TODO:
    pass

def export_godot_data_table(table : DataTable, setting : dict):
    '''
    DataTableをGodotのGDScriptファイルとして出力
    '''
    field_setting = setting["Godot"]["fields"]
    path_setting = setting["Godot"]["path"]
    replace_field_map = {}
    for key in field_setting:
        type_str = key
        code_type_str = field_setting[key]
        replace_field_map[type_str] = code_type_str

    #for record in setting.records:
    #    type_str = record.get_value("Type")
    #    code_type_str = record.get_value("CodeTypeName_Godot")
    #    replace_field_map[type_str] = code_type_str

    txt_fields = ""
    for i in range(len(table.fields)):
        field = table.fields[i]
        txt_fields += get_godot_field_str(field, replace_field_map)
    
    replace_map = {
        "DATA_NAME":table.name,
        "FIELDS":txt_fields
    }

    output_dir = path_setting["CodeOutputPath"]
    if not output_dir:
        output_dir = ""
    output_dir = OUTPUT_DIR + "/" + output_dir
    os.makedirs(output_dir, exist_ok=True)
    template_file = path_setting["TemplatePath"]
    output_file = output_dir + "/" + table.name + "Manager" + ".gd"
    export_from_template(template_file, output_file, replace_map)

def get_godot_field_str(field : DataField, replace_field_map):
    '''
    Field1つ分のテキストを作成
    '''
    txt = ""
    
    fieldtype = ""
    if field.type in replace_field_map:
        fieldtype = replace_field_map[field.type]
    else:
        fieldtype = field.type

    txt += "var " + field.name + ":"+ fieldtype + ""
    txt += LINE_STR
    return txt

def export_enum_lua(table: DataTable, setting : dict):
    '''
    Enumの出力
    '''

    path_setting = setting["Lua"]["path"]

    enums = {}
    for record in table.records:
        enum_header = record.get_value("Header")
        if enum_header == None or enum_header == "":
            # データ登録
            data_enum = enums[enum_type]
            enum_name = record.get_value("Name")
            enum_value = record.get_value("Value")
            enum_text = record.get_value("Text")
            data_enum.add_data(enum_name,enum_value,enum_text)
        else:
            enum_type = record.get_value("EnumTypeName")
            enum_type_comment = record.get_value("EnumTypeNameComment")
            enum_size= record.get_value("Size")
            #print(str(enum_type) + "," + str(enum_type_comment) + "," + str(enum_size))
            # ヘッダ登録
            enums[enum_type] = DataEnum(enum_type, enum_type_comment, enum_size)
                

    txt_enums = ""
    for key in enums.keys():
        data_enum = enums[key]
        txt_enums += get_lua_enum_str(data_enum)

    replace_map = {
        "ENUMS":txt_enums
    }

    output_dir = path_setting["EnumOutputPath"]
    if not output_dir:
        output_dir = ""
    output_dir = OUTPUT_DIR + "/" + output_dir
    os.makedirs(output_dir, exist_ok=True)
    template_file = path_setting["EnumTemplatePath"]
    output_file = output_dir + "/" + "MasterDefines" + ".lua"
    export_from_template(template_file, output_file, replace_map)

def get_lua_enum_str(data_enum : DataEnum):
    '''
    Enum1つ分のテキストを作成
    '''
    txt = ""
    txt += "-- " + data_enum.type_comment + LINE_STR
    txt += data_enum.type + " = { " + LINE_STR
    for value in data_enum.data:
        txt += "	" + value.name + " = " + str(value.value)  + "," + " -- " + value.text + LINE_STR
    txt += "}" + LINE_STR
    txt += LINE_STR
    return txt

def export_from_template(template_file, output_file, replaceMap):
    '''
    templateファイルをコピーして文字列置換したものを出力.
    '''
    txt = ""
    with open(template_file, 'r') as f:
        txt = f.read()

    for key in replaceMap:
        txt = re.sub('\\${' + key + '}', replaceMap[key], txt)
    
    with open(output_file,'w',encoding='UTF-8') as f:
        f.write(txt)

def copy_file(input, output):
    '''
    ファイルコピー
    '''
    shutil.copyfile(input, output)

def get_input_files():
    '''
    input配下のファイルを全て取得
    '''
    files = os.listdir(INPUT_DIR)
    files_file = [f for f in files if os.path.isfile(os.path.join(INPUT_DIR, f))]
    ret = []
    for f in files_file:
        if "~" in f:
            continue
        ret.append(INPUT_DIR + "/" + f)
    return ret

def load_setting(path):
    '''
    Settingファイルの読み込み
    '''
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data

def main():
    print("=================" + "load setting" + " start" "=================")
    setting = load_setting("setting.json")
    print(str(setting))
    print("=================" + "load setting" + " end" "=================")

    files = get_input_files()
    for file_path in files:
        print("=================" + file_path + " start" "=================")
        if re.search("\\.xlsx", file_path) == None:
            continue
        tables = load_and_analys_all(file_path)
        # UE4
        if setting["UE"]["enable"]:
            output_csv = setting["UE"]["enable_csv"]
            output_enum = setting["UE"]["enable_enum"]
            output_code = setting["UE"]["enable_code"]
            for table in tables:
                if table.isEnum:
                    if output_enum:
                        export_enum_ue4_header(table, setting)
                else:
                    if output_csv:
                        export_csv(table, setting)
                    if output_code:
                        export_ue4_data_table(table, setting)
        # lua
        if setting["Lua"]["enable"]:
            output_enum = setting["Lua"]["enable_enum"]
            output_data = setting["Lua"]["enable_data"]
            for table in tables:
                if table.isEnum:
                    if output_enum:
                        export_enum_lua(table, setting)
                else:
                    if output_data:
                        export_lua(table, setting)
        # godot
        if setting["Godot"]["enable"]:
            output_csv = setting["Godot"]["enable_csv"]
            output_enum = setting["Godot"]["enable_enum"]
            output_data = setting["Godot"]["enable_data"]
            output_code = setting["Godot"]["enable_code"]
            for table in tables:
                if table.isEnum:
                    export_enum_godot_header(table, setting)
                else:
                    if output_csv:
                        export_csv(table, setting)
                    if output_data:
                        # シート名にcsv_がついていたらcsvを出力
                        if ".csv" in table.name:
                            export_csv(table, setting)
                        else:
                            export_godot_dic(table, setting)
                    if output_code:
                        export_godot_data_table(table, setting)

        print("=================" + file_path + " end" "=================")

main()